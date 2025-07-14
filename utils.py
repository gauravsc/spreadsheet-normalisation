"""
Utility functions for spreadsheet processing and data type detection.
"""

import re
from typing import Dict, List, Tuple, Any, Optional
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border
import pandas as pd
import numpy as np


def load_spreadsheet(file_path: str) -> Tuple[pd.DataFrame, Dict]:
    """
    Load spreadsheet and extract both data and formatting information.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Tuple of (dataframe, formatting_info)
    """
    # Load with openpyxl to get formatting
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Get data as dataframe
    df = pd.read_excel(file_path)
    
    # Extract formatting information
    formatting_info = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_address = cell.coordinate
                formatting_info[cell_address] = {
                    'font_bold': cell.font.bold if cell.font else False,
                    'font_italic': cell.font.italic if cell.font else False,
                    'background_color': cell.fill.start_color.rgb if cell.fill.start_color.rgb != '00000000' else None,
                    'border': _extract_border_info(cell.border) if cell.border else None,
                    'number_format': cell.number_format if cell.number_format else 'General'
                }
    
    return df, formatting_info


def _extract_border_info(border) -> Dict:
    """Extract border information from cell."""
    border_info = {}
    for side in ['top', 'bottom', 'left', 'right']:
        border_side = getattr(border, side)
        if border_side and border_side.style:
            border_info[side] = border_side.style
    return border_info


def detect_data_type(value: Any, number_format: str = 'General') -> str:
    """
    Rule-based data type recognizer based on SPREADSHEETLLM paper.
    
    Args:
        value: Cell value
        number_format: Excel number format string
        
    Returns:
        Data type string
    """
    if value is None or pd.isna(value):
        return 'Empty'
    
    # Check number format strings first
    if number_format != 'General':
        if 'yyyy' in number_format or 'dd' in number_format or 'mm' in number_format:
            return 'Date'
        elif 'hh' in number_format or 'ss' in number_format:
            return 'Time'
        elif '%' in number_format:
            return 'Percentage'
        elif '$' in number_format or '€' in number_format or '£' in number_format:
            return 'Currency'
        elif '0.00E+00' in number_format or '0.00E-00' in number_format:
            return 'Scientific'
    
    # Rule-based detection
    if isinstance(value, (int, float)):
        if isinstance(value, int):
            return 'Integer'
        else:
            return 'Float'
    
    if isinstance(value, str):
        # Email detection
        if '@' in value and '.' in value:
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if re.match(email_pattern, value):
                return 'Email'
        
        # Date detection
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',
            r'\d{2}/\d{2}/\d{4}',
            r'\d{2}-\d{2}-\d{4}'
        ]
        for pattern in date_patterns:
            if re.match(pattern, value):
                return 'Date'
        
        # Time detection
        time_patterns = [
            r'\d{2}:\d{2}:\d{2}',
            r'\d{2}:\d{2}'
        ]
        for pattern in time_patterns:
            if re.match(pattern, value):
                return 'Time'
        
        # Percentage detection
        if '%' in value:
            return 'Percentage'
        
        # Currency detection
        currency_symbols = ['$', '€', '£', '¥', '₹']
        if any(symbol in value for symbol in currency_symbols):
            return 'Currency'
    
    return 'Text'


def get_cell_address(row_idx: int, col_idx: int) -> str:
    """Convert row and column indices to Excel cell address."""
    col_letter = ''
    while col_idx >= 0:
        col_letter = chr(65 + (col_idx % 26)) + col_letter
        col_idx = col_idx // 26 - 1
    return f"{col_letter}{row_idx + 1}"


def find_structural_anchors(df: pd.DataFrame, k: int = 3) -> Tuple[List[int], List[int]]:
    """
    Find structural anchors (heterogeneous rows/columns) based on SPREADSHEETLLM paper.
    
    Args:
        df: DataFrame to analyze
        k: Distance threshold for filtering - rows/columns more than k cells away from any anchor are discarded
        
    Returns:
        Tuple of (anchor_rows, anchor_columns)
    """
    anchor_rows = []
    anchor_columns = []
    
    # Find heterogeneous rows (rows with mixed data types)
    for i in range(len(df)):
        row_types = []
        for j in range(len(df.columns)):
            cell_value = df.iloc[i, j]
            cell_type = detect_data_type(cell_value)
            row_types.append(cell_type)
        
        # Check if row is heterogeneous (has multiple data types)
        unique_types = set(row_types)
        if len(unique_types) > 1:
            anchor_rows.append(i)
    
    # Find heterogeneous columns
    for j in range(len(df.columns)):
        col_types = []
        for i in range(len(df)):
            cell_value = df.iloc[i, j]
            cell_type = detect_data_type(cell_value)
            col_types.append(cell_type)
        
        unique_types = set(col_types)
        if len(unique_types) > 1:
            anchor_columns.append(j)
    
    # Apply k-neighborhood filtering: keep only rows/columns within k cells of any anchor
    if anchor_rows:
        # Find all rows within k cells of any anchor row
        all_nearby_rows = set()
        for anchor_row in anchor_rows:
            for i in range(max(0, anchor_row - k), min(len(df), anchor_row + k + 1)):
                all_nearby_rows.add(i)
        filtered_rows = sorted(list(all_nearby_rows))
    else:
        filtered_rows = []
    
    if anchor_columns:
        # Find all columns within k cells of any anchor column
        all_nearby_cols = set()
        for anchor_col in anchor_columns:
            for j in range(max(0, anchor_col - k), min(len(df.columns), anchor_col + k + 1)):
                all_nearby_cols.add(j)
        filtered_cols = sorted(list(all_nearby_cols))
    else:
        filtered_cols = []
    
    return filtered_rows, filtered_cols


def create_address_map(anchor_rows: List[int], anchor_columns: List[int]) -> Dict[str, str]:
    """
    Create address mapping for structural anchors.
    
    Args:
        anchor_rows: List of anchor row indices
        anchor_columns: List of anchor column indices
        
    Returns:
        Dictionary mapping original addresses to new addresses
    """
    address_map = {}
    
    # Map anchor rows
    for i, row_idx in enumerate(anchor_rows):
        for j in range(len(anchor_columns)):
            old_addr = get_cell_address(row_idx, anchor_columns[j])
            new_addr = get_cell_address(i, j)
            address_map[old_addr] = new_addr
    
    return address_map


def format_cell_for_markdown(address: str, value: Any, format_info: Dict = None) -> str:
    """
    Format cell information in markdown-like representation.
    
    Args:
        address: Cell address
        value: Cell value
        format_info: Formatting information
        
    Returns:
        Markdown-formatted cell string
    """
    if value is None or pd.isna(value):
        value_str = "None"
    else:
        value_str = str(value)
    
    format_str = ""
    if format_info:
        if format_info.get('font_bold'):
            format_str += "bold,"
        if format_info.get('font_italic'):
            format_str += "italic,"
        if format_info.get('background_color'):
            format_str += f"bg:{format_info['background_color']},"
        if format_info.get('number_format'):
            format_str += f"format:{format_info['number_format']},"
        
        if format_str:
            format_str = format_str.rstrip(',')
    
    return f"|{address},{value_str},{format_str}|"


def save_normalized_spreadsheet(df: pd.DataFrame, output_path: str, formatting_info: Dict = None):
    """
    Save normalized spreadsheet to Excel file.
    
    Args:
        df: Normalized DataFrame
        output_path: Output file path
        formatting_info: Optional formatting information to apply
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Normalized')
        
        # Apply formatting if provided
        if formatting_info:
            workbook = writer.book
            worksheet = writer.sheets['Normalized']
            
            for address, format_data in formatting_info.items():
                if address in worksheet:
                    cell = worksheet[address]
                    
                    if format_data.get('font_bold'):
                        cell.font = Font(bold=True)
                    if format_data.get('background_color'):
                        cell.fill = PatternFill(start_color=format_data['background_color'], 
                                              end_color=format_data['background_color'], 
                                              fill_type='solid') 
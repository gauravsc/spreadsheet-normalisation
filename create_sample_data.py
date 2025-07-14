#!/usr/bin/env python3
"""
Create sample Excel data for testing the spreadsheet normalization tool.
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def create_sample_spreadsheet(output_file: str = "sample_data.xlsx"):
    """
    Create a sample Excel spreadsheet with various data types and formatting issues.
    """
    
    # Create sample data with various issues
    data = {
        'Customer ID': [f"CUST-{i:03d}" for i in range(1, 101)],
        'Customer Name': [f"Customer {i}" for i in range(1, 101)],
        'Email': [f"customer{i}@example.com" for i in range(1, 101)],
        'Phone Number': [f"+1-555-{random.randint(100, 999)}-{random.randint(1000, 9999)}" for _ in range(100)],
        'Registration Date': [(datetime.now() - timedelta(days=random.randint(1, 365))) for _ in range(100)],
        'Last Purchase': [(datetime.now() - timedelta(days=random.randint(1, 30))) for _ in range(100)],
        'Total Spent': [round(random.uniform(100, 5000), 2) for _ in range(100)],
        'Discount Rate': [f"{random.randint(0, 25)}%" for _ in range(100)],
        'Status': random.choices(['Active', 'Inactive', 'Pending', 'Suspended'], k=100),
        'Notes': [f"Note for customer {i}" if i % 5 == 0 else "" for i in range(1, 101)]
    }
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Add some formatting issues
    df.loc[10, 'Customer Name'] = '  customer 10  '  # Extra spaces
    df.loc[15, 'Email'] = 'customer15@EXAMPLE.COM'  # Mixed case
    df.loc[20, 'Phone Number'] = '(555) 123-4567'  # Different format
    df.loc[25, 'Total Spent'] = '$1,234.56'  # Currency format
    df.loc[30, 'Discount Rate'] = '0.15'  # Decimal instead of percentage
    
    # Add some missing values
    df.loc[5, 'Email'] = None
    df.loc[12, 'Phone Number'] = ''
    df.loc[18, 'Last Purchase'] = None
    
    # Add some duplicates
    df.loc[50] = df.loc[1]  # Duplicate row
    df.loc[51] = df.loc[2]  # Another duplicate
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Data"
    
    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Format specific columns
    for row in range(2, len(df) + 2):
        # Currency formatting for Total Spent
        ws.cell(row=row, column=7).number_format = '$#,##0.00'
        
        # Percentage formatting for Discount Rate
        ws.cell(row=row, column=8).number_format = '0%'
        
        # Date formatting
        ws.cell(row=row, column=5).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=6).number_format = 'yyyy-mm-dd'
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border
    
    # Save the file
    wb.save(output_file)
    print(f"Sample spreadsheet created: {output_file}")
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    
    return output_file


def create_complex_sample(output_file: str = "complex_sample.xlsx"):
    """
    Create a more complex sample with multiple sheets and various data issues.
    """
    
    # Sheet 1: Sales Data
    sales_data = {
        'Order ID': [f"ORD-{i:04d}" for i in range(1, 201)],
        'Customer Name': [f"Customer {random.randint(1, 50)}" for _ in range(200)],
        'Product': random.choices(['Laptop', 'Phone', 'Tablet', 'Monitor', 'Keyboard'], k=200),
        'Quantity': [random.randint(1, 10) for _ in range(200)],
        'Unit Price': [round(random.uniform(100, 2000), 2) for _ in range(200)],
        'Total Amount': [],
        'Order Date': [(datetime.now() - timedelta(days=random.randint(1, 365))) for _ in range(200)],
        'Sales Rep': [f"Rep {random.randint(1, 10)}" for _ in range(200)],
        'Region': random.choices(['North', 'South', 'East', 'West'], k=200),
        'Status': random.choices(['Completed', 'Pending', 'Cancelled'], k=200)
    }
    
    # Calculate total amounts
    for i in range(200):
        total = sales_data['Quantity'][i] * sales_data['Unit Price'][i]
        sales_data['Total Amount'].append(total)
    
    sales_df = pd.DataFrame(sales_data)
    
    # Sheet 2: Inventory
    inventory_data = {
        'Product ID': [f"PROD-{i:03d}" for i in range(1, 51)],
        'Product Name': [f"Product {i}" for i in range(1, 51)],
        'Category': random.choices(['Electronics', 'Clothing', 'Books', 'Home'], k=50),
        'Stock Level': [random.randint(0, 100) for _ in range(50)],
        'Reorder Level': [random.randint(5, 20) for _ in range(50)],
        'Unit Cost': [round(random.uniform(50, 500), 2) for _ in range(50)],
        'Supplier': [f"Supplier {random.randint(1, 10)}" for _ in range(50)],
        'Last Updated': [(datetime.now() - timedelta(days=random.randint(1, 30))) for _ in range(50)]
    }
    
    inventory_df = pd.DataFrame(inventory_data)
    
    # Create Excel file with multiple sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        sales_df.to_excel(writer, sheet_name='Sales', index=False)
        inventory_df.to_excel(writer, sheet_name='Inventory', index=False)
    
    print(f"Complex sample created: {output_file}")
    print(f"Sales sheet: {sales_df.shape}")
    print(f"Inventory sheet: {inventory_df.shape}")
    
    return output_file


if __name__ == "__main__":
    print("Creating sample Excel files for testing...")
    
    # Create simple sample
    simple_file = create_sample_spreadsheet("sample_data.xlsx")
    
    # Create complex sample
    complex_file = create_complex_sample("complex_sample.xlsx")
    
    print("\nSample files created successfully!")
    print("You can now test the normalization tool with:")
    print(f"python main.py normalize --input {simple_file} --output normalized_simple.xlsx --verbose")
    print(f"python main.py analyze --input {simple_file} --output analysis.txt") 